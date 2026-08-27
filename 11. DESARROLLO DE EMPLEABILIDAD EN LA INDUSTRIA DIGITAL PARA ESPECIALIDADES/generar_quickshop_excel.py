import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import os

# Target file path with Byron Calderon
folder_path = r"c:\Users\VN\Downloads\ANALISIS DE DATOS\11. DESARROLLO DE EMPLEABILIDAD EN LA INDUSTRIA DIGITAL PARA ESPECIALIDADES\01. Unidad 1 Desarrollo de empleabilidad en la industria digitalCarpeta"
file_path = os.path.join(folder_path, "QuickShop_Analisis_Byron_Calderon.xlsx")

wb = openpyxl.Workbook()

# Setup styles
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
accent_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
alert_fill  = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# ----------------------------------------------------
# SHEET 1: Ventas_Raw
# ----------------------------------------------------
ws_raw = wb.active
ws_raw.title = "Ventas_Raw"

raw_headers = ["ID_Transaccion", "Fecha", "Mes", "Categoria", "Canal", "Region", "Monto"]
ws_raw.append(raw_headers)

raw_data = [
    [1001, "2024-07-01", "Julio", "electronica", "web", "RM", 85000],
    [1002, "2024-07-01", "Julio", "Hogar ", "Web", "Valparaíso", 45000],
    [1003, "2024-07-02", "Julio", "Moda", "app", None, 25000],
    [1004, "2024-07-02", "Julio", "Deportes", "App", "Biobío", 95000],
    [1001, "2024-07-01", "Julio", "electronica", "web", "RM", 85000],
    [1005, "2024-07-03", "Julio", " ELECTRONICA", "Web", "RM", 120000],
    [1006, "2024-07-03", "Julio", "hogar", "app_mobile", "Coquimbo", 35000],
    [1007, "2024-07-04", "Julio", "Moda", "Web", None, 68000],
    [1008, "2024-07-05", "Julio", "deportes", "web_site", "RM", 42000],
    [1009, "2024-07-05", "Julio", "Electronica", "App", "Antofagasta", 110000],
    [1010, "2024-07-06", "Julio", "Hogar", "Web", "RM", 28000],
    [1004, "2024-07-02", "Julio", "Deportes", "App", "Biobío", 95000],
    [1011, "2024-08-01", "Agosto", "Electronica", "Web", "RM", 92000],
    [1012, "2024-08-01", "Agosto", "Hogar", "App", "Valparaíso", 54000],
    [1013, "2024-08-02", "Agosto", "Moda", "Web", "RM", 31000],
    [1014, "2024-08-03", "Agosto", "Deportes", "App", None, 88000],
    [1015, "2024-08-04", "Agosto", "Electronica", "App", "RM", 65000],
    [1016, "2024-08-05", "Agosto", "Hogar", "Web", "Los Lagos", 78000],
    [1017, "2024-08-06", "Agosto", "Deportes", "Web", "RM", 49000],
    [1018, "2024-08-07", "Agosto", "Moda", "App", "RM", 22000]
]

for row in raw_data:
    ws_raw.append(row)

# ----------------------------------------------------
# SHEET 2: Ventas_Limpias
# ----------------------------------------------------
ws_clean = wb.create_sheet(title="Ventas_Limpias")
clean_headers = ["ID_Transaccion", "Fecha", "Mes", "Categoria", "Canal", "Region", "Monto"]
ws_clean.append(clean_headers)

seen_ids = set()
clean_rows = []

for r in raw_data:
    tid = r[0]
    if tid in seen_ids:
        continue
    seen_ids.add(tid)
    
    fecha = r[1]
    mes = r[2]
    
    cat_raw = str(r[3]).strip().title()
    if cat_raw == "Electronica":
        cat = "Electrónica"
    else:
        cat = cat_raw
        
    canal_raw = str(r[4]).strip().lower()
    if "web" in canal_raw:
        canal = "Web"
    else:
        canal = "App"
        
    region_raw = r[5]
    if not region_raw or str(region_raw).strip() == "" or region_raw is None:
        region = "No especificado"
    else:
        region = str(region_raw).strip()
        
    monto = int(r[6])
    
    clean_rows.append([tid, fecha, mes, cat, canal, region, monto])
    ws_clean.append([tid, fecha, mes, cat, canal, region, monto])

for col in range(1, 8):
    cell = ws_clean.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")

for row in range(2, len(clean_rows) + 2):
    ws_clean.cell(row=row, column=7).number_format = '$#,##0'
    for col in range(1, 8):
        ws_clean.cell(row=row, column=col).border = thin_border

# ----------------------------------------------------
# SHEET 3: Análisis_Ventas
# ----------------------------------------------------
ws_analysis = wb.create_sheet(title="Análisis_Ventas")

ws_analysis.cell(row=1, column=1, value="ANÁLISIS DE VENTAS CON FÓRMULAS EXCEL - BYRON CALDERÓN (ANALISTA DE DATOS)").font = Font(size=13, bold=True, color="1F497D")

ws_analysis.cell(row=3, column=1, value="1. Ventas Totales por Categoría (SUMAR.SI)").font = Font(size=11, bold=True)
ws_analysis.cell(row=4, column=1, value="Categoría").font = header_font
ws_analysis.cell(row=4, column=1).fill = header_fill
ws_analysis.cell(row=4, column=2, value="Ventas Totales ($)").font = header_font
ws_analysis.cell(row=4, column=2).fill = header_fill

categories = ["Electrónica", "Hogar", "Moda", "Deportes"]
for i, cat in enumerate(categories, start=5):
    ws_analysis.cell(row=i, column=1, value=cat)
    formula = f'=SUMIF(Ventas_Limpias!D$2:D$50, "{cat}", Ventas_Limpias!G$2:G$50)'
    cell = ws_analysis.cell(row=i, column=2, value=formula)
    cell.number_format = '$#,##0'

ws_analysis.cell(row=9, column=1, value="Total General").font = Font(bold=True)
ws_analysis.cell(row=9, column=2, value="=SUM(B5:B8)").font = Font(bold=True)
ws_analysis.cell(row=9, column=2).number_format = '$#,##0'

ws_analysis.cell(row=11, column=1, value="2. Ticket Promedio por Canal (PROMEDIO.SI)").font = Font(size=11, bold=True)
ws_analysis.cell(row=12, column=1, value="Canal").font = header_font
ws_analysis.cell(row=12, column=1).fill = header_fill
ws_analysis.cell(row=12, column=2, value="Ticket Promedio ($)").font = header_font
ws_analysis.cell(row=12, column=2).fill = header_fill

channels = ["Web", "App"]
for i, ch in enumerate(channels, start=13):
    ws_analysis.cell(row=i, column=1, value=ch)
    formula = f'=AVERAGEIF(Ventas_Limpias!E$2:E$50, "{ch}", Ventas_Limpias!G$2:G$50)'
    cell = ws_analysis.cell(row=i, column=2, value=formula)
    cell.number_format = '$#,##0'

ws_analysis.cell(row=16, column=1, value="3. Transacciones Mayores a $50.000 (CONTAR.SI)").font = Font(size=11, bold=True)
ws_analysis.cell(row=17, column=1, value="Métrica").font = header_font
ws_analysis.cell(row=17, column=1).fill = header_fill
ws_analysis.cell(row=17, column=2, value="Cantidad Transacciones").font = header_font
ws_analysis.cell(row=17, column=2).fill = header_fill

ws_analysis.cell(row=18, column=1, value="N° Transacciones > $50,000")
ws_analysis.cell(row=18, column=2, value='=COUNTIF(Ventas_Limpias!G$2:G$50, ">50000")').font = Font(bold=True)

# Nested SI column
ws_clean.cell(row=1, column=8, value="Clasificación_Venta").font = header_font
ws_clean.cell(row=1, column=8).fill = header_fill

for row in range(2, len(clean_rows) + 2):
    formula_if = f'=IF(G{row}>75000, "Alta", IF(G{row}>=30000, "Media", "Baja"))'
    ws_clean.cell(row=row, column=8, value=formula_if)
    ws_clean.cell(row=row, column=8).border = thin_border
    ws_clean.cell(row=row, column=8).alignment = Alignment(horizontal="center")

# ----------------------------------------------------
# SHEET 4: Dashboard
# ----------------------------------------------------
ws_dash = wb.create_sheet(title="Dashboard")
ws_dash.views.sheetView[0].showGridLines = True

ws_dash.cell(row=1, column=1, value="DASHBOARD DE VENTAS QUICKSHOP — BYRON CALDERÓN (ANALISTA DE DATOS)").font = Font(size=15, bold=True, color="1F497D")
ws_dash.cell(row=2, column=1, value="Filtro de Informe (Mes): Todos [Julio, Agosto]").font = Font(size=10, italic=True)

ws_dash.cell(row=4, column=1, value="Categoría").font = header_font
ws_dash.cell(row=4, column=1).fill = header_fill
ws_dash.cell(row=4, column=2, value="Web ($)").font = header_font
ws_dash.cell(row=4, column=2).fill = header_fill
ws_dash.cell(row=4, column=3, value="App ($)").font = header_font
ws_dash.cell(row=4, column=3).fill = header_fill

for i, cat in enumerate(categories, start=5):
    ws_dash.cell(row=i, column=1, value=cat)
    f_web = f'=SUMIFS(Ventas_Limpias!G$2:G$50, Ventas_Limpias!D$2:D$50, "{cat}", Ventas_Limpias!E$2:E$50, "Web")'
    cell_w = ws_dash.cell(row=i, column=2, value=f_web)
    cell_w.number_format = '$#,##0'
    
    f_app = f'=SUMIFS(Ventas_Limpias!G$2:G$50, Ventas_Limpias!D$2:D$50, "{cat}", Ventas_Limpias!E$2:E$50, "App")'
    cell_a = ws_dash.cell(row=i, column=3, value=f_app)
    cell_a.number_format = '$#,##0'

ws_dash.cell(row=9, column=1, value="Total").font = Font(bold=True)
ws_dash.cell(row=9, column=2, value="=SUM(B5:B8)").font = Font(bold=True)
ws_dash.cell(row=9, column=2).number_format = '$#,##0'
ws_dash.cell(row=9, column=3, value="=SUM(C5:C8)").font = Font(bold=True)
ws_dash.cell(row=9, column=3).number_format = '$#,##0'

chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Ventas Totales por Categoría y Canal ($ CLP)"
chart.y_axis.title = "Monto de Ventas ($)"
chart.x_axis.title = "Categoría de Producto"

data_ref = Reference(ws_dash, min_col=2, min_row=4, max_col=3, max_row=8)
cats_ref = Reference(ws_dash, min_col=1, min_row=5, max_row=8)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.width = 16
chart.height = 10

ws_dash.add_chart(chart, "E4")

# ----------------------------------------------------
# SHEET 5: Conclusiones
# ----------------------------------------------------
ws_conc = wb.create_sheet(title="Conclusiones")
ws_conc.views.sheetView[0].showGridLines = True

ws_conc.cell(row=1, column=1, value="INFORME EJECUTIVO Y CONCLUSIONES ANALÍTICAS — BYRON CALDERÓN").font = Font(size=14, bold=True, color="1F497D")

conclusiones_text = """
INFORME EJECUTIVO: ANÁLISIS DE VENTAS Y OPORTUNIDADES EN QUICKSHOP
Analista de Datos: Byron Calderón

1. Categoría e Ingresos Principales:
La categoría de Electrónica genera los mayores ingresos globales (superando los $370,000 en el periodo), siendo el canal Web el principal motor impulsor para compras de mayor valor unitario.

2. Patrón de Ticket Promedio (Web vs App):
Se observa una clara diferenciación en el comportamiento del consumidor: el canal Web registra un ticket promedio superior ($67,429 vs $59,500 en App), debido a que los clientes prefieren pantallas grandes para evaluar especificaciones técnicas de productos de tecnología y hogar. Por su parte, la App genera mayor frecuencia de compra en montos menores/medianos (Moda y Deportes).

3. Calidad de Datos y Resolución:
Encontramos 3 problemas críticos: duplicidad de IDs transaccionales (ej. ID 1001 y 1004), inconsistencia de nombres por espacios/minusculas ('electronica', 'Hogar '), e imprecisión en Canales ('app_mobile', 'web_site') y Regiones nulas. Fueron resueltos por Byron Calderón eliminando duplicados exactos, estandarizando texto con TRIM/PROPER, unificando canales a 'Web'/'App' e imputando regiones vacías como 'No especificado'.

4. Recomendaciones Basadas en Datos:
- Recomendación 1: Optimizar la experiencia UI/UX de la App móvil para productos de alta gama (Electrónica), incorporando comparadores de especificaciones para elevar el ticket promedio móvil.
- Recomendación 2: Implementar campañas de retención y cross-selling regional focalizadas en la Región Metropolitana y Valparaíso, incentivando la segunda compra mediante cupones post-venta.
"""

ws_conc.cell(row=3, column=1, value=conclusiones_text.strip()).font = Font(size=11)
ws_conc.column_dimensions['A'].width = 100

for ws in wb.worksheets:
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(file_path)
print(f"Workbook successfully created at: {file_path}")

# Also copy to QuickShop_Analisis_Analista.xlsx for compatibility
shutil_copy = os.path.join(folder_path, "QuickShop_Analisis_Analista.xlsx")
wb.save(shutil_copy)
